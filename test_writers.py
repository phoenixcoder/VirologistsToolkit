from unittest import TestCase
from unittest.mock import Mock, MagicMock, patch, PropertyMock, ANY
from structures import SubstitutionMatrix
from writers import SubstitutionMatrixDataWriter
import openpyxl

class TestNormalizedBiasDataWriter(TestCase):

    @patch('structures.SubstitutionMatrix')
    @patch('openpyxl.worksheet.worksheet.Worksheet')
    @patch('openpyxl.Workbook')
    def testWrite(self, mockWorkbook, mockWorksheet,
        mockSubstitutionMatrix):
        mockBase = Mock()
        mockWorkbook.create_sheet = Mock(return_value = mockWorksheet)

        testSubjectHeader = "Test Header"
        testSubject = "Subject"
        testInvalidSubject = "Invalid Subject"
        testValidSubjects = {testSubject: mockSubstitutionMatrix}
        testInvalidSubjects = [testInvalidSubject]

        testWriter = SubstitutionMatrixDataWriter(testSubjectHeader,
            testValidSubjects, testInvalidSubjects, mockBase)
        testWriter._writeHeaders = Mock()
        testWriter._writeSubject = Mock()
        testWriter.write(mockWorkbook)

        mockWorkbook.create_sheet.assert_called_once_with(title =
            "Results")
        testWriter._writeHeaders.assert_called_once_with(mockWorksheet)
        testWriter._writeSubject.assert_called_once_with(testSubject, 
            mockSubstitutionMatrix, mockWorksheet)

    @patch('structures.SubstitutionMatrix')
    @patch('openpyxl.worksheet.worksheet.Worksheet')
    def test_writeHeaders(self, mockWorksheet, mockSubstitutionMatrix):
        mockBase = MagicMock()
        mockWorksheet.cell = MagicMock(spec=openpyxl.cell.cell)
        pValueMock = PropertyMock()
        type(mockWorksheet.cell(row = ANY, column = ANY)) \
            .value = pValueMock

        testSubjectHeader = "Test Header"
        testSubject = "Subject"
        testInvalidSubject = "Invalid Subject"
        testValidSubjects = {testSubject: mockSubstitutionMatrix}
        testInvalidSubjects = [testInvalidSubject]
        mockBase.__iter__.return_value = [1, 2,]

        testWriter = SubstitutionMatrixDataWriter(testSubjectHeader,
            testValidSubjects, testInvalidSubjects, mockBase)
        testWriter._writeHeaders(mockWorksheet)


        self.assertEquals(mockBase.__iter__.call_count, 3,
            "Nucleobase iterator must be called three times.")
        mockWorksheet.cell.assert_any_call(row = 1, column = 1)
        mockWorksheet.cell.assert_any_call(row = 1, column = 2)

        pValueMock.assert_any_call("Virus")
        pValueMock.assert_any_call("1 -> 2")
        pValueMock.assert_any_call("2 -> 1")

    @patch('openpyxl.worksheet.worksheet.Worksheet')
    @patch('structures.SubstitutionMatrix')
    def test_writeSubject(self, mockSubstitutionMatrix, mockWorksheet):
        testSubjectHeader = "Test Header"
        testSubject = "Subject"
        testInvalidSubject = "Invalid Subject"
        testValidSubjects = {testSubject: mockSubstitutionMatrix}
        testInvalidSubjects = [testInvalidSubject]
        testSubjectName = "Test Subject Name"
        mockBase = MagicMock()
        mockBase.__iter__.return_value = [0, 1]
        mockSubstitutionMatrix.getCopy.return_value = [[1, 2], [3, 4]]
        pValueMock = PropertyMock()
        type(mockWorksheet.cell(row = ANY, column = ANY)).value \
            = pValueMock

        testWriter = SubstitutionMatrixDataWriter(testSubjectHeader,
            testValidSubjects, testInvalidSubjects, mockBase)
        testWriter._writeSubject(testSubjectName,
            mockSubstitutionMatrix, mockWorksheet)

        pValueMock.assert_any_call(testSubjectName)
        pValueMock.assert_any_call(2)
        pValueMock.assert_any_call(3)
