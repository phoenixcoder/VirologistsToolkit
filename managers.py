from openpyxl import load_workbook, Workbook
from structures import SubstitutionMatrix, DNA
from readers import ObservedExpectedMatricesReader

class XlManager:
    """Manages all conversions of Excel data to necessary structures.

    Reads raw data into substitution matrices after validation.  Writes
    a dictionary of values to an Excel sheet.
    """

    def readResultsFromWorkbook(self, inputFilepath, reader):
        """Retrieves a dictionary of valid observed/expected matrices.

        Uses the name of the worksheet as the key, if the sheet has both
        a valid observed/expected matrix.  The result is that the class'
        matrix dictionary is populated with all valid entries and/or the
        invalid sheets list will be populated with the name of all the
        worksheets where validation failed.

        Retrieve the results of this process via the reader's instance
        variables.  An agent is normally using this class so the agent
        will know which field to call on when this is done processing.
        """
        if inputFilepath != None:
            workbook = load_workbook(inputFilepath)
            wsNames = workbook.get_sheet_names()
            for name in wsNames:
                worksheet = workbook.get_sheet_by_name(name)
                reader.read(worksheet)

    def writeResultsToWorkbook(self, outputFilepath, writer):
        """Creates a workbook to write results to.

        Delegates control of the worksheets to the writer as the writer
        has knowledge of the structure of the data set. This facilitates
        writing data to the file and does nothing more.

        :param outputFilepath: Path to the file
        :param writer:         Writer of the data
        """
        workbook = Workbook()
        writer.write(workbook)
