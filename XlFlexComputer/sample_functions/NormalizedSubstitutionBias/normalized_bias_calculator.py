from openpyxl import load_workbook, Workbook
from enum import IntEnum
from math import fsum, pow, sqrt

# Configuration
INPUT = "data/2SubstitutionAnlysML_INPUT.xlsx"
OUTPUT = "data/SubstitutionAnlysResults.xlsx"
resultsWsName = "Results"

# Observed and Expected Matrices Row and Column Starting Indices
omStartRi = 18
omStartCi = 1
emStartRi = 18
emStartCi = 7

class DNA(IntEnum):

    """DNA Bases Representing the Position on the Row or Column Headers

    The positions proceed from left-to-right on the row header and from
    top to bottom on the column header.
    """

    A = 0
    C = 1
    G = 2
    T = 3

    def __str__(self):
        return self.name

def validateHeaders(startRi, startCi, worksheet):
    """Ensures the matrix we're looking at has the correct headers.

    The matrix must look like the following to be valid:
    _ |  A   C   G   T
    -------------------
    A |  -   N   N   N
      |
    C |  N   -   N   N
      |
    G |  N   N   -   N
      |
    T |  N   N   N   -
    **Notes: There are no gaps between columns and rows.  They are
    there for aesthetic purposes.

    [ Legend ]
     * Hyphens indicate '-', characters, or blanks.
     * N indicates numeral, numeric strings, or blanks.

    :param startRi:     Row index to start validating at.
    :param startCi:     Column index to start validating at.
    :param worksheet:   Worksheet to do validation on.
    """
    result = True
    members = DNA.__members__
    headerValues = list(members.keys())
    headerValues.append(None)

    # Validate on the row and column headers at the same time.
    for index in range(len(headerValues)):
        currRowValue = worksheet.cell(row = startRi + index,
            column = startCi).value
        currColValue = worksheet.cell(row = startRi,
            column = startCi + index).value

        if currRowValue == currColValue and currRowValue in headerValues:
            try:
                headerValues.remove(currRowValue)
            except ValueError:
                result = False
                break
        else:
            result = False
            break

    return result

def validWorksheet(worksheet):
    """Runs all validation methods to validate the worksheet."""
    result = validateHeaders(omStartRi, omStartCi, worksheet) \
        and validateHeaders(emStartRi, emStartCi, worksheet)
    return result

def extractAndCalculateData(worksheet, formula):
    """Extracts the data and performs the calculations.

    Applies the normalized bias formula substitution-by-substitution.

    :param worksheet: Worksheet the formula is to be applied to.
    :param formula:   Formula that is to be applied to the worksheet.
    :returns:         Dictionary-of-dictionaries with float values that were
                      pulled from the worksheet.
    """
    resultData = dict()
    for row in DNA:
        for col in DNA:
            if row != col:
                oValRi = omStartRi + int(row) + 1
                oValCi = omStartCi + int(col) + 1
                eValRi = emStartRi + int(row) + 1
                eValCi = emStartCi + int(col) + 1
                oVal = worksheet \
                    .cell(row = oValRi, column = oValCi).value
                eVal = worksheet \
                    .cell(row = eValRi, column = eValCi).value
                amount = convertToFloat(formula(oVal, eVal))
                resultData[str(row) + " -> " + str(col)] = amount

    return resultData

def calculateStandardScore(extractedData):
    """Calculates the Standard Score on the given dataset.

    For each substitution difference (labeled as diff), the formula is:

    (standard score) = (diff - mean) / (standard deviation)

    :params extractedData: Dictionary-of-Dictionaries with float values.
    :returns:              Standard score for each datapoint.
    """
    results = dict()
    mean = calculateMean(extractedData)
    stdDev = calculateStandardDeviation(extractedData, mean)

    if stdDev != 0:
        for title in extractedData.keys():
            subResults = dict()
            for subTitle in extractedData[title].keys():
                subResults[subTitle] = \
                    (extractedData[title][subTitle] - mean) / stdDev
            results[title] = subResults

    print("Mean: " + str(mean))
    print("Standard Deviation: " + str(stdDev))
    return results

def biasFormula(observed, expected):
    """Performs the normalized bias calculation."""
    result = -2.0
    floatObs = convertToFloat(observed)
    floatExp = convertToFloat(expected)
    if (floatObs + floatExp) != 0:
        result = (floatObs - floatExp) / (floatObs**2 + floatExp**2)
    return result

def differenceFormula(observed, expected):
    """Performs a simple substraction calculation between observed/expected."""
    floatObs = convertToFloat(observed)
    floatExp = convertToFloat(expected)
    return convertToFloat(abs(floatObs - floatExp))

def convertToFloat(target):
    result = 0.0
    if target is not None:
        try:
            result = float(target)
        except ValueError:
            result = 0.0

    return result

def calculateMean(extractedWorksheetData):
    """Calculates the mean of the float data set.

    Assumes the extractedWorksheetData is a dictionary-of-dictionaries with the
    second level of dictionaries containing all the same keys.

    :params extractedWorksheetData: Dictionary-of-dictionaries containing float
                                    values.
    :returns:                       Mean of all the float values.
    """
    results = 0.0
    meanNumerator = 0.0
    numItems = 0.0
    for valSet in extractedWorksheetData.values():
        values = valSet.values()
        numItems = numItems + len(values)
        meanNumerator = meanNumerator + fsum(values)
    
    if numItems != 0:
        results = convertToFloat(meanNumerator / numItems)

    return results

def calculateStandardDeviation(extractedWorksheetData, mean):
    """Calculates the standard deviation of the float data set.

    Assumes the extractedWorksheetData is a dictionary-of-dictionaries with the
    second level of dictionaries containing all the same keys.

    :params extractedWorksheetData: Dictionary-of-dictionaries containing
                                    float values.
    :params mean:                   Mean over all the float values.
    :returns:                       Standard deviation of all the float values.
    """
    results = 0.0
    numItems = 0.0
    stdNumerator = 0.0
    for valSet in extractedWorksheetData.values():
        values = valSet.values()
        numItems = numItems + len(values)
        for val in values:
            stdNumerator = stdNumerator + pow(val - mean, 2)

    if numItems != 0.0:
        results = sqrt(stdNumerator / numItems)

    return results

def printExtractedData(extractedWorksheetData, title):
    """For testing purposes only, prints all the keys and subkeys of the dict.

    Prints out the keys and subkeys of the dictionary-of-dictionaries, which is
    the extracted worksheet data.
    """
    print(title)
    print("========================")
    for worksheetTitle in extractedWorksheetData.keys():
        print(worksheetTitle)
        worksheet = extractedWorksheetData[worksheetTitle]
        for subTitle in worksheet.keys():
            print(subTitle + ": " + str(worksheet[subTitle]))

# SCRIPT START

inWorkbook = load_workbook(INPUT, data_only = True)
outWorkbook = Workbook()

wsResults = outWorkbook.active
wsResults.title = resultsWsName

results = dict()
failures = []
wsNames = inWorkbook.get_sheet_names()
# Extract the data
for name in wsNames:
    worksheet = inWorkbook.get_sheet_by_name(name)
    if validWorksheet(worksheet):
        data = extractAndCalculateData(worksheet, differenceFormula)
        results[worksheet.title] = data
    else:
        failures.append(worksheet.title)

results = calculateStandardScore(results)

# Results Printing
headerRi = 1
headerCi = 1
wsResults.cell(row = headerRi, column = headerCi).value = "Virus"

# Print names of successfully completed worksheets and prints the data to the
# results worksheet.
print("Successfully Processed Worksheets\n" +
      "====================================")
currOutRi = headerRi + 1
headerPrinted = False
for key in results.keys():
    wsResults.cell(row = currOutRi, column = headerCi).value = key

    data = results[key]
    for ci, substitutionName in enumerate(data.keys()):
        if currOutRi == 2:
            wsResults.cell(row = headerRi, column = headerCi + ci + 1) \
                .value = substitutionName
        wsResults.cell(row = currOutRi, column = headerCi + ci + 1) \
            .value = float(data[substitutionName])

    currOutRi = currOutRi + 1
    print(key)

# Print erred sheet names
print("\nFailed Worksheets\n" + 
      "=================")
for name in failures:
    print(name)

outWorkbook.save(filename = OUTPUT)
